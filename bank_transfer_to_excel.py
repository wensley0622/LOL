#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
银行转账记录整理脚本
功能：从PDF银行流水中提取 日期、金额、对方账户名，输出为Excel表格

使用方法：
  1. 把PDF文件放在同一目录下
  2. 运行：python3 bank_transfer_to_excel.py 你的文件.pdf
  3. 会在同目录生成 你的文件_转账记录.xlsx

依赖安装（只需第一次运行前装一次）：
  pip3 install pdfplumber openpyxl
"""

import sys
import os
import re
from datetime import datetime

try:
    import pdfplumber
except ImportError:
    print("❌ 缺少依赖，请先运行：pip3 install pdfplumber openpyxl")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("❌ 缺少依赖，请先运行：pip3 install pdfplumber openpyxl")
    sys.exit(1)


# ============================================================
# 常见银行转账记录的日期格式（可自行添加）
# ============================================================
DATE_PATTERNS = [
    r'\d{4}[-/年]\d{1,2}[-/月]\d{1,2}日?',   # 2024-01-15 / 2024/01/15 / 2024年01月15日
    r'\d{2}[-/]\d{1,2}[-/]\d{1,2}',           # 24-01-15 / 01-15-24
]

# 金额格式：支持逗号分隔和带正负号
AMOUNT_PATTERN = r'[+-]?\d{1,3}(,\d{3})*(\.\d{2})'


def extract_from_tables(pdf):
    """方法1：尝试从PDF中提取表格数据（银行流水通常是表格形式）"""
    rows = []
    for page in pdf.pages:
        tables = page.extract_tables()
        for table in tables:
            if not table:
                continue
            for row in table:
                if row and any(cell and str(cell).strip() for cell in row):
                    # 把 None 替换为空字符串，并去除首尾空格
                    cleaned = [str(cell).strip() if cell else '' for cell in row]
                    rows.append(cleaned)
    return rows


def extract_from_text(pdf):
    """方法2：从PDF纯文本中用正则提取转账信息"""
    rows = []
    for page in pdf.pages:
        text = page.extract_text()
        if not text:
            continue
        # 按行分割
        for line in text.split('\n'):
            line = line.strip()
            if not line:
                continue
            # 尝试匹配日期
            date_match = None
            for pattern in DATE_PATTERNS:
                m = re.search(pattern, line)
                if m:
                    date_match = m.group()
                    break
            # 尝试匹配金额
            amount_matches = re.findall(AMOUNT_PATTERN, line)
            # 重新匹配，保留完整金额（含逗号和负号）
            amount_full = re.findall(r'[+-]?\d{1,3}(?:,\d{3})*(?:\.\d{2})', line)

            if date_match or amount_full:
                rows.append({
                    'raw': line,
                    'date': date_match or '',
                    'amounts': amount_full,
                })
    return rows


def parse_date(date_str):
    """将各种日期格式统一为 YYYY-MM-DD"""
    if not date_str:
        return ''
    # 替换中文
    s = date_str.replace('年', '-').replace('月', '-').replace('日', '')
    s = s.replace('/', '-')
    # 尝试解析
    for fmt in ['%Y-%m-%d', '%y-%m-%d']:
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return date_str  # 解析不了就原样返回


def clean_amount(amount_str):
    """清理金额：去掉逗号，保留负号和小数"""
    if not amount_str:
        return ''
    s = amount_str.replace(',', '')
    return s


def identify_columns(header_row):
    """
    识别表头中各列的含义
    返回字典：{'日期': 列号, '金额': 列号, '对方': 列号, ...}
    """
    col_map = {}
    if not header_row:
        return col_map

    keywords = {
        '日期': ['日期', '交易日期', '记账日期', '时间', '交易时间'],
        '金额': ['金额', '交易金额', '发生额', '借方金额', '贷方金额', '收入', '支出', '借方', '贷方'],
        '余额': ['余额', '账户余额', '可用余额'],
        '对方': ['对方', '对方户名', '对方账户', '交易对方', '收款人', '付款人', '对方名称', '对手'],
        '摘要': ['摘要', '备注', '用途', '交易摘要', '交易类型', '附言'],
    }

    for i, cell in enumerate(header_row):
        cell_lower = cell.lower() if cell else ''
        for key, words in keywords.items():
            for word in words:
                if word in cell_lower:
                    col_map[key] = i
                    break

    return col_map


def process_table_rows(rows):
    """处理表格方式提取的数据，整理成统一格式"""
    if not rows:
        return []

    # 尝试识别表头（通常是第一行或前两行）
    col_map = identify_columns(rows[0])
    if not col_map:
        # 尝试第二行
        if len(rows) > 1:
            col_map = identify_columns(rows[1])
            if col_map:
                rows = rows[1:]  # 跳过第一行
    header = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []

    results = []

    # 如果成功识别了列含义
    if col_map:
        for row in data_rows:
            if len(row) < max(col_map.values()) + 1:
                continue
            entry = {}
            for key, col_idx in col_map.items():
                val = row[col_idx] if col_idx < len(row) else ''
                entry[key] = val

            # 清洗
            entry['日期'] = parse_date(entry.get('日期', ''))
            entry['金额'] = clean_amount(entry.get('金额', ''))

            # 判断收支方向
            amount_str = entry.get('金额', '')
            if amount_str.startswith('-'):
                entry['方向'] = '支出'
            elif amount_str:
                entry['方向'] = '收入'

            results.append(entry)
    else:
        # 没识别出表头，直接输出原始行
        for row in data_rows:
            results.append({'原始数据': ' | '.join(row)})

    return results


def process_text_rows(text_rows):
    """处理纯文本方式提取的数据"""
    results = []
    for item in text_rows:
        entry = {}
        entry['日期'] = parse_date(item.get('date', ''))

        amounts = item.get('amounts', [])
        if amounts:
            entry['金额'] = clean_amount(amounts[0])
            if len(amounts) > 1:
                entry['余额'] = clean_amount(amounts[1])

        # 判断收支
        amt = entry.get('金额', '')
        if amt.startswith('-'):
            entry['方向'] = '支出'
        elif amt:
            entry['方向'] = '收入'

        # 从原文中尝试提取对方名称（去掉日期和金额后的部分）
        raw = item.get('raw', '')
        remaining = raw
        if item.get('date'):
            remaining = remaining.replace(item['date'], '', 1)
        for a in amounts:
            remaining = remaining.replace(a, '', 1)
        entry['对方/摘要'] = remaining.strip()

        results.append(entry)

    return results


def write_to_excel(data, output_path):
    """将数据写入Excel文件，带格式"""
    wb = Workbook()
    ws = wb.active
    ws.title = "转账记录"

    if not data:
        print("⚠️ 没有提取到数据，请检查PDF格式")
        return

    # 确定所有列名
    all_keys = []
    for entry in data:
        for key in entry:
            if key not in all_keys:
                all_keys.append(key)

    # 优先排序列名
    priority = ['日期', '方向', '金额', '对方', '余额', '摘要', '对方/摘要']
    ordered_keys = []
    for p in priority:
        if p in all_keys:
            ordered_keys.append(p)
            all_keys.remove(p)
    ordered_keys.extend(all_keys)  # 剩余的列放后面

    # 写表头
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, size=11, color='FFFFFF')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col, key in enumerate(ordered_keys, 1):
        cell = ws.cell(row=1, column=col, value=key)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 写数据
    for row_idx, entry in enumerate(data, 2):
        for col_idx, key in enumerate(ordered_keys, 1):
            value = entry.get(key, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

            # 金额列右对齐
            if key in ('金额', '余额'):
                cell.alignment = Alignment(horizontal='right', vertical='center')

            # 支出行标红
            if key == '方向' and value == '支出':
                cell.font = Font(color='FF0000')
            elif key == '方向' and value == '收入':
                cell.font = Font(color='008000')

    # 自动调整列宽
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                # 中文字符按2个宽度计算
                length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 添加筛选
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    print(f"✅ 已生成Excel文件：{output_path}")
    print(f"   共 {len(data)} 条记录")


def main():
    if len(sys.argv) < 2:
        print("用法：python3 bank_transfer_to_excel.py <PDF文件路径>")
        print("示例：python3 bank_transfer_to_excel.py 银行流水.pdf")
        sys.exit(1)

    pdf_path = sys.argv[1]
    if not os.path.exists(pdf_path):
        print(f"❌ 找不到文件：{pdf_path}")
        sys.exit(1)

    # 生成输出文件名
    base_name = os.path.splitext(pdf_path)[0]
    output_path = f"{base_name}_转账记录.xlsx"

    print(f"📖 正在读取PDF：{pdf_path}")

    try:
        with pdfplumber.open(pdf_path) as pdf:
            print(f"   共 {len(pdf.pages)} 页")

            # 方法1：优先尝试表格提取
            print("🔍 尝试从表格提取数据...")
            table_rows = extract_from_tables(pdf)
            data = process_table_rows(table_rows)

            if data and len(data) > 0 and '原始数据' not in data[0]:
                print(f"   ✓ 表格提取成功，找到 {len(data)} 条记录")
            else:
                # 方法2：用正则从文本提取
                print("🔍 表格提取效果不佳，改用文本解析...")
                text_rows = extract_from_text(pdf)
                data = process_text_rows(text_rows)
                print(f"   ✓ 文本解析完成，找到 {len(data)} 条记录")

            write_to_excel(data, output_path)

    except Exception as e:
        print(f"❌ 处理失败：{e}")
        print("💡 提示：如果PDF是扫描件（图片形式），需要先用OCR工具转换")
        sys.exit(1)


if __name__ == '__main__':
    main()
