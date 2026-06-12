#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
银行转账记录整理脚本（招商银行APP交易详情版）

功能：从招商银行APP交易详情截图中提取转账信息，汇总到Excel表格
方案：macOS 内置 Vision 框架进行 OCR

使用方法：
  python3 bank_transfer_ocr.py 图片1.png [图片2.png ...]
  python3 bank_transfer_ocr.py *.png

依赖（只需装一次）：
  pip3 install --user openpyxl Pillow
"""

import sys
import os
import json
import tempfile
import subprocess
import re

# ============================================================
# 依赖检查
# ============================================================
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("❌ 缺少 openpyxl，请运行：pip3 install --user openpyxl")
    sys.exit(1)

try:
    from PIL import Image as PILImage
except ImportError:
    print("❌ 缺少 Pillow，请运行：pip3 install --user Pillow")
    sys.exit(1)


# ============================================================
# Swift OCR 引擎
# ============================================================
SWIFT_OCR_CODE = r'''
import Vision
import Foundation
import AppKit

guard CommandLine.arguments.count > 1 else { exit(1) }
let imagePath = CommandLine.arguments[1]
guard let image = NSImage(contentsOfFile: imagePath) else { exit(1) }
guard let cgImage = image.cgImage(forProposedRect: nil, context: nil, hints: nil) else { exit(1) }
let imgWidth = CGFloat(cgImage.width)
let imgHeight = CGFloat(cgImage.height)

let request = VNRecognizeTextRequest()
request.recognitionLevel = .accurate
request.recognitionLanguages = ["zh-Hans", "zh-Hant", "en"]
request.usesLanguageCorrection = true

let handler = VNImageRequestHandler(cgImage: cgImage, options: [:])
try! handler.perform([request])

guard let observations = request.results else { print("[]"); exit(0) }

var items: [[String: Any]] = []
for obs in observations {
    guard let candidate = obs.topCandidates(1).first else { continue }
    let bbox = obs.boundingBox
    let x = bbox.origin.x * imgWidth
    let y = (1 - bbox.origin.y - bbox.height) * imgHeight
    let w = bbox.width * imgWidth
    let h = bbox.height * imgHeight
    items.append([
        "text": candidate.string,
        "x": Double(x), "y": Double(y),
        "w": Double(w), "h": Double(h)
    ])
}
items.sort { ($0["y"] as! Double) < ($1["y"] as! Double) }
let jsonData = try! JSONSerialization.data(withJSONObject: items, options: [])
print(String(data: jsonData, encoding: .utf8)!)
'''


def ocr_image(image_path):
    """调用 Swift + Vision 进行 OCR"""
    swift_file = os.path.join(tempfile.gettempdir(), 'bank_ocr_helper.swift')
    if not os.path.exists(swift_file):
        with open(swift_file, 'w', encoding='utf-8') as f:
            f.write(SWIFT_OCR_CODE)

    try:
        result = subprocess.run(
            ['swift', swift_file, os.path.abspath(image_path)],
            capture_output=True, text=True, timeout=60
        )
    except subprocess.TimeoutExpired:
        print("  ⚠️ OCR 超时")
        return []

    if result.returncode != 0:
        print(f"  ⚠️ OCR 出错：{result.stderr[:200]}")
        return []

    try:
        return json.loads(result.stdout.strip())
    except json.JSONDecodeError:
        return []


# ============================================================
# 行分组
# ============================================================
def group_into_rows(ocr_items, y_tolerance=None):
    """按 y 坐标将 OCR 文字分组为行"""
    if not ocr_items:
        return []

    sorted_items = sorted(ocr_items, key=lambda i: i['y'] + i['h'] / 2)

    if y_tolerance is None:
        heights = [i['h'] for i in sorted_items if i['h'] > 0]
        avg_h = sum(heights) / len(heights) if heights else 20
        y_tolerance = avg_h * 0.6

    rows = []
    current_row = [sorted_items[0]]

    for item in sorted_items[1:]:
        cur_y = sum(i['y'] + i['h'] / 2 for i in current_row) / len(current_row)
        item_y = item['y'] + item['h'] / 2

        if abs(item_y - cur_y) < y_tolerance:
            current_row.append(item)
        else:
            current_row.sort(key=lambda i: i['x'])
            rows.append(current_row)
            current_row = [item]

    if current_row:
        current_row.sort(key=lambda i: i['x'])
        rows.append(current_row)

    return rows


def row_text(row):
    """拼接一行文字"""
    return ' '.join(i['text'] for i in row)


# ============================================================
# 交易详情解析 —— 招商银行APP格式
# ============================================================

# 已知字段标签（左侧标签 → Excel列名）
FIELD_LABELS = {
    '交易卡号': '交易卡号',
    '交易时间': '交易时间',
    '付款银行': '付款银行',
    '收款银行': '收款银行',
    '付款账号': '付款账号',
    '收款账号': '收款账号',
    '付款人': '付款人',
    '收款人': '收款人',
    '转账附言': '转账附言',
    '附言': '转账附言',
    '银行交易类型': '银行交易类型',
    '交易类型': '交易类型',
    '交易渠道': '交易渠道',
    '对方户名': '对方户名',
    '对方账户': '对方账户',
    '摘要': '摘要',
    '备注': '备注',
    '所属账本': '所属账本',
    '所厲账本': '所属账本',  # OCR 可能识别错 "属" 为 "厲"
}


def parse_transaction_detail(rows):
    """
    解析招商银行APP交易详情页

    页面布局（重要特征）：
    - 顶部：对方名称 + 金额（大字）
    - 中间：余额
    - 下方：键值对字段
      ★ 值在标签上方（先显示值，再显示标签）
      ★ 长文本会被拆成多行，尾部可能在标签同行右侧
    """
    record = {
        '对方名称': '',
        '金额': '',
        '收支方向': '',
        '余额': '',
        '交易卡号': '',
        '交易时间': '',
        '付款银行': '',
        '收款银行': '',
        '付款账号': '',
        '收款账号': '',
        '转账附言': '',
        '银行交易类型': '',
    }

    # APP界面干扰文字（不是有效数据）
    SKIP_TEXTS = ['薪福专区', '轻松查工资', '查看往来', '给TA转账',
                  '开电子收据', '不计入本月', '记录点什么', '请选择＞',
                  '请选择', '日 他人转入', '分类', '所厲账本', '所属账本']

    # 第一步：找金额和余额（通常在前几行，字号大）
    for row in rows:
        text = row_text(row)

        # 匹配金额：+￥5,172.41 或 -￥1,000.00
        amount_match = re.search(r'([+-])?\s*[￥¥]\s*([\d,]+\.?\d*)', text)
        if amount_match and not record['金额']:
            direction = amount_match.group(1) or ''
            amount = amount_match.group(2).replace(',', '')
            record['金额'] = amount
            if direction == '+':
                record['收支方向'] = '收入'
            elif direction == '-':
                record['收支方向'] = '支出'
            continue

        # 匹配余额
        balance_match = re.search(r'余额\s*[￥¥]?\s*([\d,]+\.?\d*)', text)
        if balance_match and not record['余额']:
            record['余额'] = balance_match.group(1).replace(',', '')
            continue

    # 第二步：找对方名称（通常在金额上方，带序号圆圈或不带）
    for row in rows:
        text = row_text(row)
        if '交易详情' in text:
            continue
        if re.search(r'[￥¥]', text):
            continue
        if re.match(r'^[①②③④⑤⑥⑦⑧⑨⑩]?\s*[一-鿿]{2,4}$', text.strip()):
            name = re.sub(r'^[①②③④⑤⑥⑦⑧⑨⑩]\s*', '', text.strip())
            if not record['对方名称'] and name:
                record['对方名称'] = name
            break

    # 第三步：解析键值对字段
    # 招商银行APP特征：值在标签上方！
    # 所以我们先用标签找位置，再往上一行取值
    label_positions = {}  # col_name → row_index

    for idx, row in enumerate(rows):
        text = row_text(row)

        for label_key, col_name in FIELD_LABELS.items():
            if label_key in text:
                label_positions[col_name] = idx

    # 根据标签位置提取值
    for col_name, label_idx in label_positions.items():
        label_row = rows[label_idx]
        label_text = row_text(label_row)

        # 跳过干扰字段
        if col_name == '所属账本':
            continue

        # 找到标签文字块的位置
        label_item = None
        for item in label_row:
            for label_key, cn in FIELD_LABELS.items():
                if cn == col_name and label_key in item['text']:
                    label_item = item
                    break
            if label_item:
                break

        value_parts = []

        # 策略1：标签同行的右侧文字（可能是长值的尾部，如"行"）
        if label_item:
            for item in label_row:
                if item['x'] > label_item['x'] + label_item['w'] * 0.5:
                    val = item['text'].strip()
                    if val and not any(kw in val for kw in SKIP_TEXTS):
                        value_parts.append(val)

        # 策略2：标签上一行的文字（值在标签上方——招商银行APP的核心特征）
        if label_idx > 0:
            prev_row = rows[label_idx - 1]
            prev_text = row_text(prev_row)
            # 确认上一行不是另一个标签行
            is_other_label = False
            for other_col, other_idx in label_positions.items():
                if other_idx == label_idx - 1 and other_col != col_name:
                    is_other_label = True
                    break

            if not is_other_label:
                # 上一行就是值
                # 如果上一行包含金额或余额，跳过（已单独处理）
                if not re.search(r'[￥¥]|余额', prev_text):
                    # 过滤掉干扰文字
                    clean_parts = []
                    for item in prev_row:
                        val = item['text'].strip()
                        if val and not any(kw in val for kw in SKIP_TEXTS):
                            clean_parts.append(val)
                    if clean_parts:
                        # 值在上方，尾部在标签右侧，拼接顺序：上方值 + 右侧尾部
                        value_parts = [' '.join(clean_parts)] + value_parts

        # 如果上方和右侧都没值，尝试下方
        if not value_parts and label_idx < len(rows) - 1:
            next_row = rows[label_idx + 1]
            next_text = row_text(next_row)
            if not any(kw in next_text for kw in SKIP_TEXTS):
                is_next_label = any(
                    label_key in next_text
                    for label_key in FIELD_LABELS
                )
                if not is_next_label:
                    clean_parts = []
                    for item in next_row:
                        val = item['text'].strip()
                        if val and not any(kw in val for kw in SKIP_TEXTS):
                            clean_parts.append(val)
                    value_parts = [' '.join(clean_parts)]

        full_value = ''.join(value_parts).strip()
        if full_value:
            record[col_name] = full_value

    # 第四步：从"银行交易类型"推断收支方向
    if not record['收支方向']:
        txn_type = record.get('银行交易类型', '')
        if '汇入' in txn_type or '转入' in txn_type or '收入' in txn_type:
            record['收支方向'] = '收入'
        elif '汇出' in txn_type or '转出' in txn_type or '支出' in txn_type:
            record['收支方向'] = '支出'

    return record


# ============================================================
# Excel 输出
# ============================================================
def write_excel(records, output_path):
    """将转账记录写入 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "转账记录"

    if not records:
        print("⚠️ 没有提取到任何记录")
        return

    # 列定义
    columns = [
        ('交易时间', 20),
        ('对方名称', 12),
        ('收支方向', 10),
        ('金额', 14),
        ('余额', 14),
        ('付款银行', 36),
        ('付款账号', 22),
        ('收款银行', 36),
        ('收款账号', 22),
        ('交易卡号', 26),
        ('转账附言', 14),
        ('银行交易类型', 14),
    ]

    # 样式
    header_font = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='C0C0C0'),
        right=Side(style='thin', color='C0C0C0'),
        top=Side(style='thin', color='C0C0C0'),
        bottom=Side(style='thin', color='C0C0C0'),
    )
    income_font = Font(color='006600', bold=True)
    expense_font = Font(color='CC0000', bold=True)
    even_fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')

    # 写表头
    for col, (name, width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[cell.column_letter].width = width

    # 写数据
    for row_idx, entry in enumerate(records, 2):
        for col_idx, (name, _) in enumerate(columns, 1):
            value = entry.get(name, '')
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center', wrap_text=True)

            # 金额列右对齐
            if name in ('金额', '余额'):
                cell.alignment = Alignment(horizontal='right', vertical='center')

            # 收支方向标色
            if name == '收支方向':
                if value == '收入':
                    cell.font = income_font
                elif value == '支出':
                    cell.font = expense_font

            # 金额标色
            if name == '金额':
                direction = entry.get('收支方向', '')
                if direction == '收入':
                    cell.font = income_font
                    cell.value = f'+{value}' if value else ''
                elif direction == '支出':
                    cell.font = expense_font
                    cell.value = f'-{value}' if value else ''

            # 交替行底色
            if row_idx % 2 == 0:
                cell.fill = even_fill

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 自动筛选
    ws.auto_filter.ref = ws.dimensions

    # 汇总行
    summary_row = len(records) + 2
    ws.cell(row=summary_row, column=1, value='汇总').font = Font(bold=True, size=12)

    income_total = 0
    expense_total = 0
    for entry in records:
        amt = entry.get('金额', '')
        direction = entry.get('收支方向', '')
        if amt:
            try:
                v = float(amt)
                if direction == '收入':
                    income_total += v
                elif direction == '支出':
                    expense_total += v
            except ValueError:
                pass

    ws.cell(row=summary_row, column=3, value='收入合计').font = Font(bold=True)
    ws.cell(row=summary_row, column=4, value=f'{income_total:,.2f}').font = Font(bold=True, color='006600')
    ws.cell(row=summary_row, column=4).alignment = Alignment(horizontal='right')

    ws.cell(row=summary_row + 1, column=3, value='支出合计').font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=4, value=f'{expense_total:,.2f}').font = Font(bold=True, color='CC0000')
    ws.cell(row=summary_row + 1, column=4).alignment = Alignment(horizontal='right')

    ws.cell(row=summary_row + 2, column=3, value='净额').font = Font(bold=True)
    net = income_total - expense_total
    ws.cell(row=summary_row + 2, column=4, value=f'{net:,.2f}')
    ws.cell(row=summary_row + 2, column=4).font = Font(bold=True, color='006600' if net >= 0 else 'CC0000')
    ws.cell(row=summary_row + 2, column=4).alignment = Alignment(horizontal='right')

    wb.save(output_path)
    print(f"\n✅ 已生成 Excel：{output_path}")
    print(f"   共 {len(records)} 条记录")
    print(f"   收入合计：{income_total:,.2f}")
    print(f"   支出合计：{expense_total:,.2f}")
    print(f"   净额：{net:,.2f}")


# ============================================================
# 主程序
# ============================================================
def main():
    if len(sys.argv) < 2:
        print("=" * 50)
        print("银行转账记录整理脚本（交易详情截图版）")
        print("=" * 50)
        print()
        print("用法：python3 bank_transfer_ocr.py <图片> [更多图片...]")
        print()
        print("示例：")
        print("  python3 bank_transfer_ocr.py 流水1.png")
        print("  python3 bank_transfer_ocr.py *.png")
        sys.exit(1)

    # 收集图片文件
    image_files = []
    for arg in sys.argv[1:]:
        if os.path.isfile(arg):
            ext = os.path.splitext(arg)[1].lower()
            if ext in ('.png', '.jpg', '.jpeg', '.tiff', '.bmp', '.heic'):
                image_files.append(arg)
            else:
                print(f"⚠️ 跳过非图片文件：{arg}")
        else:
            print(f"⚠️ 文件不存在：{arg}")

    if not image_files:
        print("❌ 没有找到有效的图片文件")
        sys.exit(1)

    # 按文件名排序
    image_files.sort()

    print(f"📖 共 {len(image_files)} 张图片待处理\n")

    all_records = []

    for i, img_path in enumerate(image_files, 1):
        print(f"🔍 [{i}/{len(image_files)}] {os.path.basename(img_path)}")

        # OCR
        items = ocr_image(img_path)
        if not items:
            print("  ⚠️ 未识别到文字，跳过")
            continue
        print(f"  ✓ 识别到 {len(items)} 个文字块")

        # 分行
        rows = group_into_rows(items)

        # 解析
        record = parse_transaction_detail(rows)
        all_records.append(record)

        # 显示识别结果
        direction = record.get('收支方向', '?')
        amount = record.get('金额', '?')
        name = record.get('对方名称', '?')
        time = record.get('交易时间', '?')
        print(f"  → {direction} {amount} | {name} | {time}")

    # 生成 Excel
    if all_records:
        first_dir = os.path.dirname(os.path.abspath(image_files[0]))
        output_path = os.path.join(first_dir, '银行转账记录.xlsx')
        write_excel(all_records, output_path)
    else:
        print("\n⚠️ 未能提取到任何记录")


if __name__ == '__main__':
    main()
